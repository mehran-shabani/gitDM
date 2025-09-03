import io
import os
from datetime import datetime, timedelta
from typing import Dict, List, Optional
from django.conf import settings
from django.utils import timezone
from django.template.loader import render_to_string
from django.core.mail import EmailMessage

# Import for PDF generation
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# Import for Excel generation
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Import for charts
import matplotlib
matplotlib.use('Agg')  # Use non-interactive backend
import matplotlib.pyplot as plt
import seaborn as sns

from encounters.models import Patient, Encounter
from laboratory.models import LabTest, LabResult
from pharmacy.models import Medication
from security.models import DoctorProfile
from .models import Report, PatientAnalytics, DoctorAnalytics, SystemAnalytics
from .services import PatientAnalyticsService, DoctorAnalyticsService, SystemAnalyticsService


class ReportGenerationService:
    """سرویس تولید گزارش‌های PDF و Excel"""
    
    def __init__(self):
        self.patient_service = PatientAnalyticsService()
        self.doctor_service = DoctorAnalyticsService()
        self.system_service = SystemAnalyticsService()
        
        # تنظیمات فونت فارسی برای PDF
        self._setup_persian_font()
    
    def _setup_persian_font(self):
        """تنظیم فونت فارسی برای ReportLab"""
        try:
            # مسیر فونت را بر اساس سیستم تنظیم کنید
            font_path = os.path.join(settings.BASE_DIR, 'static', 'fonts', 'IRANSans.ttf')
            if os.path.exists(font_path):
                pdfmetrics.registerFont(TTFont('IRANSans', font_path))
                self.persian_font = 'IRANSans'
            else:
                self.persian_font = 'Helvetica'  # Fallback to default font
        except:
            self.persian_font = 'Helvetica'
    
    def generate_report(self, report: Report) -> str:
        """تولید گزارش بر اساس نوع و فرمت"""
        try:
            report.status = 'processing'
            report.save()
            
            # تولید گزارش بر اساس نوع
            if report.report_type == 'patient_summary':
                file_path = self._generate_patient_summary(report)
            elif report.report_type == 'doctor_performance':
                file_path = self._generate_doctor_performance(report)
            elif report.report_type == 'system_overview':
                file_path = self._generate_system_overview(report)
            else:
                file_path = self._generate_custom_report(report)
            
            report.file_path = file_path
            report.status = 'completed'
            report.completed_at = timezone.now()
            report.save()
            
            return file_path
            
        except Exception as e:
            report.status = 'failed'
            report.error_message = str(e)
            report.save()
            raise
    
    def _generate_patient_summary(self, report: Report) -> str:
        """تولید گزارش خلاصه بیمار"""
        patient = report.patient
        if not patient:
            raise ValueError("بیمار برای گزارش مشخص نشده است")
        
        # جمع‌آوری داده‌ها
        data = self._collect_patient_data(patient, report.start_date, report.end_date)
        
        # تولید گزارش بر اساس فرمت
        if report.format == 'pdf':
            return self._generate_patient_pdf(patient, data, report)
        elif report.format == 'excel':
            return self._generate_patient_excel(patient, data, report)
        else:
            return self._generate_patient_csv(patient, data, report)
    
    def _collect_patient_data(self, patient: Patient, start_date: Optional[datetime.date], 
                            end_date: Optional[datetime.date]) -> Dict:
        """جمع‌آوری داده‌های بیمار"""
        if not end_date:
            end_date = timezone.now().date()
        if not start_date:
            start_date = end_date - timedelta(days=90)
        
        # مواجهات
        encounters = Encounter.objects.filter(
            patient=patient,
            encounter_date__range=[start_date, end_date]
        ).order_by('-encounter_date')
        
        # آزمایش‌ها
        lab_tests = LabTest.objects.filter(
            patient=patient,
            test_date__range=[start_date, end_date]
        ).select_related('results').order_by('-test_date')
        
        # داروها
        medications = Medication.objects.filter(
            patient=patient,
            prescribed_date__range=[start_date, end_date]
        ).order_by('-prescribed_date')
        
        # آخرین آنالیتیکس
        latest_analytics = PatientAnalytics.objects.filter(
            patient=patient
        ).order_by('-date').first()
        
        # نمودارها
        glucose_chart = self.patient_service.get_glucose_chart_data(patient, 'month')
        hba1c_chart = self.patient_service.get_hba1c_trend_data(patient)
        
        return {
            'patient': patient,
            'encounters': encounters,
            'lab_tests': lab_tests,
            'medications': medications,
            'analytics': latest_analytics,
            'glucose_chart': glucose_chart,
            'hba1c_chart': hba1c_chart,
            'start_date': start_date,
            'end_date': end_date
        }
    
    def _generate_patient_pdf(self, patient: Patient, data: Dict, report: Report) -> str:
        """تولید PDF برای گزارش بیمار"""
        # ایجاد مسیر فایل
        filename = f"patient_report_{patient.id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        file_path = os.path.join(settings.MEDIA_ROOT, 'reports', filename)
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        
        # ایجاد PDF
        doc = SimpleDocTemplate(file_path, pagesize=A4)
        story = []
        styles = getSampleStyleSheet()
        
        # سبک‌های سفارشی
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#2c3e50'),
            spaceAfter=30,
            alignment=TA_CENTER,
            fontName=self.persian_font
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#34495e'),
            spaceAfter=12,
            fontName=self.persian_font
        )
        
        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontSize=11,
            fontName=self.persian_font
        )
        
        # عنوان گزارش
        story.append(Paragraph(f"گزارش پزشکی بیمار", title_style))
        story.append(Spacer(1, 0.2*inch))
        
        # اطلاعات بیمار
        patient_info = [
            ['نام و نام خانوادگی:', patient.full_name],
            ['کد ملی:', patient.national_id],
            ['تاریخ تولد:', patient.date_of_birth.strftime('%Y-%m-%d') if patient.date_of_birth else '-'],
            ['جنسیت:', patient.get_gender_display()],
            ['پزشک معالج:', patient.primary_doctor.user.get_full_name() if patient.primary_doctor else '-']
        ]
        
        patient_table = Table(patient_info, colWidths=[2*inch, 4*inch])
        patient_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), self.persian_font),
            ('FONTSIZE', (0, 0), (-1, -1), 11),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#7f8c8d')),
            ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        
        story.append(patient_table)
        story.append(Spacer(1, 0.3*inch))
        
        # خلاصه آماری
        if data['analytics']:
            story.append(Paragraph("خلاصه وضعیت", heading_style))
            
            summary_data = [
                ['شاخص', 'مقدار', 'وضعیت'],
                ['میانگین قند خون', f"{data['analytics'].avg_glucose:.1f} mg/dL" if data['analytics'].avg_glucose else '-', 
                 self._get_glucose_status(data['analytics'].avg_glucose)],
                ['میانگین HbA1c', f"{data['analytics'].avg_hba1c:.1f}%" if data['analytics'].avg_hba1c else '-',
                 self._get_hba1c_status(data['analytics'].avg_hba1c)],
                ['تعداد ویزیت‌ها', str(data['analytics'].encounters_count), '-'],
                ['امتیاز پایبندی', f"{data['analytics'].compliance_score:.0f}%" if data['analytics'].compliance_score else '-',
                 self._get_compliance_status(data['analytics'].compliance_score)]
            ]
            
            summary_table = Table(summary_data, colWidths=[2*inch, 2*inch, 2*inch])
            summary_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), self.persian_font),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498db')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#bdc3c7')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#ecf0f1')]),
            ]))
            
            story.append(summary_table)
            story.append(Spacer(1, 0.3*inch))
        
        # نمودار قند خون
        if report.metadata.get('include_charts', True):
            story.append(Paragraph("روند قند خون", heading_style))
            
            # تولید نمودار
            chart_path = self._create_glucose_chart(data['glucose_chart'], patient.id)
            if chart_path and os.path.exists(chart_path):
                img = Image(chart_path, width=6*inch, height=3*inch)
                story.append(img)
                story.append(Spacer(1, 0.2*inch))
        
        # جدول آزمایش‌های اخیر
        story.append(PageBreak())
        story.append(Paragraph("آزمایش‌های اخیر", heading_style))
        
        lab_data = [['تاریخ', 'نوع آزمایش', 'نتیجه', 'واحد', 'وضعیت']]
        for test in data['lab_tests'][:10]:  # حداکثر 10 آزمایش اخیر
            result = test.results.first()
            if result:
                lab_data.append([
                    test.test_date.strftime('%Y-%m-%d'),
                    test.get_test_type_display(),
                    result.value,
                    result.unit or '-',
                    'نرمال' if result.is_normal else 'غیرنرمال'
                ])
        
        if len(lab_data) > 1:
            lab_table = Table(lab_data, colWidths=[1.5*inch, 2*inch, 1*inch, 1*inch, 1.5*inch])
            lab_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), self.persian_font),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2ecc71')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#bdc3c7')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#ecf0f1')]),
            ]))
            story.append(lab_table)
        
        # ایجاد PDF
        doc.build(story)
        
        return file_path
    
    def _generate_patient_excel(self, patient: Patient, data: Dict, report: Report) -> str:
        """تولید Excel برای گزارش بیمار"""
        filename = f"patient_report_{patient.id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        file_path = os.path.join(settings.MEDIA_ROOT, 'reports', filename)
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        
        # ایجاد Excel writer
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            # برگه اطلاعات بیمار
            patient_info = {
                'عنوان': ['نام و نام خانوادگی', 'کد ملی', 'تاریخ تولد', 'جنسیت', 'پزشک معالج'],
                'مقدار': [
                    patient.full_name,
                    patient.national_id,
                    patient.date_of_birth.strftime('%Y-%m-%d') if patient.date_of_birth else '-',
                    patient.get_gender_display(),
                    patient.primary_doctor.user.get_full_name() if patient.primary_doctor else '-'
                ]
            }
            df_info = pd.DataFrame(patient_info)
            df_info.to_excel(writer, sheet_name='اطلاعات بیمار', index=False)
            
            # برگه آزمایش‌ها
            lab_records = []
            for test in data['lab_tests']:
                result = test.results.first()
                if result:
                    lab_records.append({
                        'تاریخ': test.test_date.strftime('%Y-%m-%d'),
                        'نوع آزمایش': test.get_test_type_display(),
                        'نتیجه': result.value,
                        'واحد': result.unit or '-',
                        'محدوده نرمال': result.reference_range or '-',
                        'وضعیت': 'نرمال' if result.is_normal else 'غیرنرمال'
                    })
            
            if lab_records:
                df_labs = pd.DataFrame(lab_records)
                df_labs.to_excel(writer, sheet_name='آزمایش‌ها', index=False)
            
            # برگه داروها
            med_records = []
            for med in data['medications']:
                med_records.append({
                    'تاریخ تجویز': med.prescribed_date.strftime('%Y-%m-%d'),
                    'نام دارو': med.medication_name,
                    'دوز': med.dosage,
                    'تناوب': med.get_frequency_display(),
                    'مدت': f"{med.duration} روز" if med.duration else '-',
                    'وضعیت': med.get_status_display()
                })
            
            if med_records:
                df_meds = pd.DataFrame(med_records)
                df_meds.to_excel(writer, sheet_name='داروها', index=False)
            
            # برگه خلاصه آماری
            if data['analytics']:
                summary_data = {
                    'شاخص': [
                        'میانگین قند خون',
                        'حداقل قند خون',
                        'حداکثر قند خون',
                        'میانگین HbA1c',
                        'تعداد ویزیت‌ها',
                        'تعداد آزمایش‌ها',
                        'امتیاز پایبندی'
                    ],
                    'مقدار': [
                        f"{data['analytics'].avg_glucose:.1f}" if data['analytics'].avg_glucose else '-',
                        f"{data['analytics'].min_glucose:.1f}" if data['analytics'].min_glucose else '-',
                        f"{data['analytics'].max_glucose:.1f}" if data['analytics'].max_glucose else '-',
                        f"{data['analytics'].avg_hba1c:.1f}" if data['analytics'].avg_hba1c else '-',
                        data['analytics'].encounters_count,
                        data['analytics'].lab_tests_count,
                        f"{data['analytics'].compliance_score:.0f}" if data['analytics'].compliance_score else '-'
                    ]
                }
                df_summary = pd.DataFrame(summary_data)
                df_summary.to_excel(writer, sheet_name='خلاصه آماری', index=False)
            
            # قالب‌بندی
            workbook = writer.book
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                
                # تنظیم عرض ستون‌ها
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # قالب‌بندی هدر
                for cell in worksheet[1]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
        
        return file_path
    
    def _generate_doctor_performance(self, report: Report) -> str:
        """تولید گزارش عملکرد پزشک"""
        doctor = report.doctor
        if not doctor:
            raise ValueError("پزشک برای گزارش مشخص نشده است")
        
        # جمع‌آوری داده‌ها
        data = self._collect_doctor_data(doctor, report.start_date, report.end_date)
        
        # تولید گزارش بر اساس فرمت
        if report.format == 'pdf':
            return self._generate_doctor_pdf(doctor, data, report)
        elif report.format == 'excel':
            return self._generate_doctor_excel(doctor, data, report)
        else:
            return self._generate_doctor_csv(doctor, data, report)
    
    def _collect_doctor_data(self, doctor: DoctorProfile, start_date: Optional[datetime.date],
                           end_date: Optional[datetime.date]) -> Dict:
        """جمع‌آوری داده‌های عملکرد پزشک"""
        if not end_date:
            end_date = timezone.now().date()
        if not start_date:
            start_date = end_date - timedelta(days=30)
        
        # آنالیتیکس پزشک
        latest_analytics = DoctorAnalytics.objects.filter(
            doctor=doctor
        ).order_by('-date').first()
        
        # لیست بیماران
        patients = Patient.objects.filter(primary_doctor=doctor)
        
        # ویزیت‌های دوره
        encounters = Encounter.objects.filter(
            doctor=doctor,
            encounter_date__range=[start_date, end_date]
        ).select_related('patient')
        
        # نمودار توزیع بیماران
        distribution_chart = self.doctor_service.get_patient_distribution_data(doctor)
        
        return {
            'doctor': doctor,
            'analytics': latest_analytics,
            'patients': patients,
            'encounters': encounters,
            'distribution_chart': distribution_chart,
            'start_date': start_date,
            'end_date': end_date
        }
    
    def _generate_system_overview(self, report: Report) -> str:
        """تولید گزارش نمای کلی سیستم"""
        # جمع‌آوری داده‌ها
        data = self._collect_system_data(report.start_date, report.end_date)
        
        # تولید گزارش بر اساس فرمت
        if report.format == 'pdf':
            return self._generate_system_pdf(data, report)
        elif report.format == 'excel':
            return self._generate_system_excel(data, report)
        else:
            return self._generate_system_csv(data, report)
    
    def _collect_system_data(self, start_date: Optional[datetime.date],
                           end_date: Optional[datetime.date]) -> Dict:
        """جمع‌آوری داده‌های سیستم"""
        if not end_date:
            end_date = timezone.now().date()
        if not start_date:
            start_date = end_date - timedelta(days=30)
        
        # آنالیتیکس سیستم
        latest_analytics = SystemAnalytics.objects.order_by('-date').first()
        
        # داده‌های نمای کلی
        overview_data = self.system_service.get_system_overview_data()
        
        # نمودارهای روند
        user_trend = self.system_service.get_trend_chart_data('users', 'month')
        patient_trend = self.system_service.get_trend_chart_data('patients', 'month')
        
        return {
            'analytics': latest_analytics,
            'overview': overview_data,
            'user_trend': user_trend,
            'patient_trend': patient_trend,
            'start_date': start_date,
            'end_date': end_date
        }
    
    def _create_glucose_chart(self, chart_data: Dict, patient_id: int) -> str:
        """ایجاد نمودار قند خون"""
        try:
            plt.figure(figsize=(10, 6))
            
            # استخراج داده‌ها
            for dataset in chart_data['datasets']:
                x_values = [d['x'] for d in dataset['data']]
                y_values = [d['y'] for d in dataset['data']]
                
                plt.plot(x_values, y_values, marker='o', label=dataset['label'])
            
            plt.xlabel('تاریخ')
            plt.ylabel('قند خون (mg/dL)')
            plt.title('روند قند خون')
            plt.legend()
            plt.grid(True, alpha=0.3)
            plt.xticks(rotation=45)
            plt.tight_layout()
            
            # ذخیره نمودار
            chart_path = os.path.join(settings.MEDIA_ROOT, 'charts', f'glucose_{patient_id}_{datetime.now().timestamp()}.png')
            os.makedirs(os.path.dirname(chart_path), exist_ok=True)
            plt.savefig(chart_path, dpi=150, bbox_inches='tight')
            plt.close()
            
            return chart_path
        except Exception as e:
            print(f"Error creating chart: {e}")
            return None
    
    def _get_glucose_status(self, value: Optional[float]) -> str:
        """تعیین وضعیت قند خون"""
        if not value:
            return '-'
        if value < 70:
            return 'پایین'
        elif value <= 180:
            return 'نرمال'
        else:
            return 'بالا'
    
    def _get_hba1c_status(self, value: Optional[float]) -> str:
        """تعیین وضعیت HbA1c"""
        if not value:
            return '-'
        if value < 7:
            return 'عالی'
        elif value < 8:
            return 'خوب'
        elif value < 9:
            return 'متوسط'
        else:
            return 'نیاز به بهبود'
    
    def _get_compliance_status(self, value: Optional[float]) -> str:
        """تعیین وضعیت پایبندی"""
        if not value:
            return '-'
        if value >= 80:
            return 'عالی'
        elif value >= 60:
            return 'خوب'
        elif value >= 40:
            return 'متوسط'
        else:
            return 'ضعیف'
    
    def schedule_report(self, report_config: Dict) -> None:
        """زمان‌بندی تولید گزارش‌های دوره‌ای"""
        # این متد می‌تواند با Celery پیاده‌سازی شود
        pass
    
    def send_report_email(self, report: Report, recipients: List[str]) -> None:
        """ارسال گزارش از طریق ایمیل"""
        if report.status != 'completed' or not report.file_path:
            raise ValueError("گزارش هنوز آماده نیست")
        
        subject = f"گزارش {report.get_report_type_display()} - {report.created_at.strftime('%Y-%m-%d')}"
        
        message = f"""
        سلام،
        
        گزارش درخواستی شما آماده است.
        
        نوع گزارش: {report.get_report_type_display()}
        تاریخ تولید: {report.created_at.strftime('%Y-%m-%d %H:%M')}
        
        گزارش به پیوست ارسال شده است.
        
        با تشکر،
        سیستم مدیریت دیابت GitDM
        """
        
        email = EmailMessage(
            subject=subject,
            body=message,
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=recipients
        )
        
        # پیوست فایل گزارش
        if os.path.exists(report.file_path):
            with open(report.file_path, 'rb') as f:
                email.attach_file(report.file_path)
        
        email.send()